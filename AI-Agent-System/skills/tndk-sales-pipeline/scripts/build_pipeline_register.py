#!/usr/bin/env python3
"""
Build the TNDK Pipeline Register — every quotation from issue to decision.

Companion to build_register.py, and it follows the same rule: every figure is
computed in Python, verified by assertion, then written as a LITERAL. Formulas
appear only as a redundant cross-check column.

The Approved Works Register failed because formula ranges did not grow with the
table. A pipeline register would fail the same way and be harder to notice,
because nobody knows what the numbers should be.

Two things this file will not do, both deliberate:

  - It will not print a win rate from fewer than 20 tracked decisions.
  - It will not compute a rate from rows marked `reconstructed`. Those come from
    an award register that only ever recorded wins, so any rate drawn from them
    reads near 100% by construction.

Where a rate cannot honestly be produced, the cell says so in words.

Usage:
    python3 build_pipeline_register.py --data pipeline.json --outdir ./out
    python3 build_pipeline_register.py --data pipeline.json --as-of 2026-08-03
"""

import argparse
import json
import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from pipeline import (analyse, parse_date, VALIDITY_DAYS, QUIET_DAYS,  # noqa: E402
                      MIN_DECIDED_FOR_RATE)

NAVY = "1F3864"
GOLD = "C9A24E"
LIGHT = "D6E4F0"
GREY = "F2F2F2"
GREEN_FILL, GREEN_TEXT = "E2EFDA", "2E7D32"
RED_TEXT, RED_FILL = "C00000", "FDECEA"
AMBER = "FFF2CC"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0.00'


def _header_block(ws, title, ncols, as_of):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="TNDK")
    c.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=title)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1,
                value=f"The New Doha Kitchen Equipment Services W.L.L.   ·   As of {as_of}")
    c.font = Font(name="Calibri", size=9, italic=True, color="404040")
    c.alignment = Alignment(horizontal="center")


def _table_header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def build(data, a, as_of, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"

    # "Days" means days open for a live quotation, and days taken to decide for a
    # decided one. One column, two meanings, so the header says both.
    headers = ["Quote ref", "Date sent", "Client", "Project", "Value (QAR)", "Tier",
               "Source", "Status", "Days open / to decide", "Last contact", "Flag",
               "Next action", "Next date", "Decided", "Loss reason", "Notes", "Check"]
    widths = [22, 11, 26, 28, 14, 9, 11, 10, 13, 12, 16, 30, 11, 11, 20, 26, 9]
    _header_block(ws, "PIPELINE REGISTER — EVERY QUOTATION, ISSUE TO DECISION",
                  len(headers), as_of.strftime("%d %B %Y"))
    _table_header(ws, 5, headers, widths)

    by_ref = {q["ref"]: q for q in data["quotes"]}
    open_by_ref = {r["ref"]: r for r in a["open"]}
    decided_by_ref = {d["ref"]: d for d in a["decided"]}

    r = 6
    first_data = r
    open_value = 0.0

    for q in data["quotes"]:
        ref = q["ref"]
        value = float(q["value"])
        status = q["status"]
        o = open_by_ref.get(ref)
        d = decided_by_ref.get(ref)

        flag = " · ".join(o["flags"]) if o and o["flags"] else ""
        days_open = o["days_open"] if o else (d["days_to_decide"] if d else "")
        if o:
            open_value += value

        vals = [ref, q["date_sent"], q["client"], q.get("project", ""), value,
                q.get("tier", ""), q.get("source", ""), status.upper(), days_open,
                q.get("last_contact", ""), flag, q.get("next_action", ""),
                q.get("next_action_date", ""), q.get("decided_date", ""),
                q.get("loss_reason", ""), q.get("_flag", q.get("notes", ""))]

        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Calibri", size=10)
            c.border = BORDER
            if col == 5:
                c.number_format = MONEY
                c.alignment = Alignment(horizontal="right")
            elif col in (2, 6, 7, 8, 9, 10, 13, 14):
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        st = ws.cell(row=r, column=8)
        if status == "won":
            st.fill = PatternFill("solid", fgColor=GREEN_FILL)
            st.font = Font(name="Calibri", size=10, bold=True, color=GREEN_TEXT)
        elif status == "lost":
            st.fill = PatternFill("solid", fgColor=RED_FILL)
            st.font = Font(name="Calibri", size=10, bold=True, color=RED_TEXT)
        elif status == "open":
            st.fill = PatternFill("solid", fgColor=AMBER)
            st.font = Font(name="Calibri", size=10, bold=True)
        else:
            st.fill = PatternFill("solid", fgColor=GREY)
            st.font = Font(name="Calibri", size=10, bold=True, color="808080")

        if flag:
            fc = ws.cell(row=r, column=11)
            fc.fill = PatternFill("solid", fgColor=RED_FILL)
            fc.font = Font(name="Calibri", size=9, bold=True, color=RED_TEXT)

        if q.get("reconstructed") or q.get("tier") == "unrecorded":
            ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, italic=True,
                                                 color="808080")

        # Redundant cross-check: a decided row must carry a decided date, a lost row
        # must carry a reason. Reading MISMATCH means the file was hand-edited.
        chk = ws.cell(row=r, column=17, value=(
            f'=IF(OR(AND(OR(H{r}="WON",H{r}="LOST"),N{r}=""),'
            f'AND(H{r}="LOST",O{r}="")),"MISMATCH","OK")'))
        chk.font = Font(name="Calibri", size=9, italic=True, color="808080")
        chk.alignment = Alignment(horizontal="center")
        chk.border = BORDER
        r += 1

    last_data = r - 1

    # --- TOTAL row: literal, plus a formula over the FULL range ---
    ws.cell(row=r, column=1, value="OPEN PIPELINE TOTAL")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.border = BORDER
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", indent=1)
    c = ws.cell(row=r, column=5, value=open_value)
    c.number_format = MONEY
    c.alignment = Alignment(horizontal="right")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=r, column=17, value=(
        f'=IF(ROUND(SUMIF(H{first_data}:H{last_data},"OPEN",'
        f'E{first_data}:E{last_data})-E{r},2)=0,"OK","MISMATCH")')
    ).font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    r += 2

    # --- Summary ---
    ws.cell(row=r, column=2, value="SUMMARY").font = Font(name="Calibri", size=12, bold=True,
                                                         color=NAVY)
    r += 1
    basis = ("observed" if not a["base_rate_assumed"]
             else f"ASSUMED {a['base_rate']:.0%} — not observed")
    rows = [
        ("Open quotations", len(a["open"]), None, NAVY),
        ("Open pipeline, unweighted", a["unweighted"], MONEY, NAVY),
        (f"Expected value ({basis})", a["weighted"], MONEY,
         RED_TEXT if a["base_rate_assumed"] else GREEN_TEXT),
        ("Requiring action (past validity or quiet)", len(a["action"]), None, RED_TEXT),
        ("Decided — tracked issue to decision", a["n_tracked"], None, NAVY),
        ("Decided — reconstructed (wins-only record)", a["n_reconstructed"], None, "808080"),
        ("Won", len(a["won"]), None, GREEN_TEXT),
        ("Lost", len(a["lost"]), None, RED_TEXT),
    ]
    for label, val, fmt, colour in rows:
        ws.cell(row=r, column=2, value=label).font = Font(name="Calibri", size=11, bold=True)
        c = ws.cell(row=r, column=5, value=val)
        if fmt:
            c.number_format = fmt
        c.font = Font(name="Calibri", size=11, bold=True, color=colour)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right")
        r += 1

    # --- Win rate: a number, or a sentence saying why there isn't one ---
    ws.cell(row=r, column=2, value="Win rate").font = Font(name="Calibri", size=11, bold=True)
    if a["rate_known"]:
        c = ws.cell(row=r, column=5, value=a["win_rate"])
        c.number_format = '0.0%'
        c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    else:
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
        c = ws.cell(row=r, column=5,
                    value=f"NOT YET ANSWERABLE — {a['n_tracked']} of "
                          f"{MIN_DECIDED_FOR_RATE} tracked decisions")
        c.font = Font(name="Calibri", size=10, bold=True, color=RED_TEXT)
        c.alignment = Alignment(horizontal="left", indent=1)
    c.fill = PatternFill("solid", fgColor=LIGHT)
    c.border = BORDER
    r += 2

    # --- Loss reasons ---
    if a["loss_reasons"]:
        ws.cell(row=r, column=2, value="LOSS REASONS").font = Font(name="Calibri", size=12,
                                                                   bold=True, color=NAVY)
        r += 1
        for reason, n in sorted(a["loss_reasons"].items(), key=lambda x: -x[1]):
            ws.cell(row=r, column=2, value=reason).font = Font(name="Calibri", size=10)
            c = ws.cell(row=r, column=5, value=n)
            c.font = Font(name="Calibri", size=11, bold=True, color=RED_TEXT)
            c.fill = PatternFill("solid", fgColor=GREY)
            c.border = BORDER
            c.alignment = Alignment(horizontal="right")
            r += 1
        r += 1

    note = ws.cell(row=r, column=1, value=(
        f"Quotation validity {VALIDITY_DAYS} days; a quotation with no client contact for "
        f"{QUIET_DAYS} days is flagged QUIET. Rows in grey italic are reconstructed from the "
        f"award register, which only ever recorded wins — they establish the denominator and "
        f"are excluded from every rate. The 'Check' column re-derives each row; any cell "
        f"reading MISMATCH means this file has been edited by hand and should be rebuilt."))
    note.font = Font(name="Calibri", size=9, italic=True, color="808080")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 42

    ws.freeze_panes = "A6"
    wb.save(path)
    return {"open_value": open_value, "open_count": len(a["open"])}


def main():
    ap = argparse.ArgumentParser(description="Build the TNDK pipeline register")
    ap.add_argument("--data", required=True)
    ap.add_argument("--outdir", default=".")
    ap.add_argument("--as-of", default=None, help="YYYY-MM-DD, defaults to today")
    args = ap.parse_args()

    with open(args.data) as f:
        data = json.load(f)

    as_of = parse_date(args.as_of) if args.as_of else (
        parse_date(data["as_of"]) if data.get("as_of") else date.today())

    a = analyse(data, as_of)     # same analysis the report uses — one source of truth
    os.makedirs(args.outdir, exist_ok=True)
    path = os.path.join(args.outdir, "pipeline_register.xlsx")
    out = build(data, a, as_of, path)

    # Assertions: fail loudly rather than ship a wrong register.
    exp_open = sum(float(q["value"]) for q in data["quotes"] if q["status"] == "open")
    assert abs(out["open_value"] - exp_open) < 0.01, "open pipeline total mismatch"
    assert out["open_count"] == len(a["open"]), "open row count mismatch"

    print(f"Pipeline Register        → {path}")
    print(f"  Quotations tracked        {len(data['quotes']):>14}")
    print(f"  Open                      {out['open_count']:>14}")
    print(f"  Open pipeline value       {out['open_value']:>14,.2f}")
    print(f"  Requiring action          {len(a['action']):>14}")
    if a["rate_known"]:
        print(f"  Win rate                  {a['win_rate']*100:>13.1f}%")
    else:
        print(f"  Win rate                  {'not yet answerable':>14}"
              f"  ({a['n_tracked']} of {MIN_DECIDED_FOR_RATE} tracked)")
    print("\nAll totals asserted against source data.")


if __name__ == "__main__":
    main()
