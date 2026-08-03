#!/usr/bin/env python3
"""
Rebuild the TNDK Approved Works Register and Amounts to Receive.

Writes COMPUTED VALUES, not formulas. The live register failed because formula
ranges did not grow with the table: its total row summed rows 1-3 only, returning
18,250 against a real book of 758,100, and every balance cell read 0.00.

A spreadsheet that must be recalculated to be correct is a spreadsheet that will
someday be read before it is recalculated.

Every figure here is computed in Python, verified by assertion, then written as a
literal. Formulas are added only as a redundant cross-check in a separate column.

Usage:
    python3 build_register.py --data awards.json --outdir ./out
"""

import argparse
import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
GOLD = "C9A24E"
LIGHT = "D6E4F0"
GREY = "F2F2F2"
GREEN_FILL, GREEN_TEXT = "E2EFDA", "2E7D32"
RED_TEXT, RED_FILL = "C00000", "FDECEA"
AMBER = "FFF2CC"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0.00'


def _header_block(ws, title, ncols, as_of):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="TNDK")
    c.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=title)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1,
                value=f"The New Doha Kitchen Equipment Services W.L.L.   ·   As of {as_of}")
    c.font = Font(name="Calibri", size=9, italic=True, color="404040")
    c.alignment = Alignment(horizontal="center")


def _table_header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def build_approved(awards, as_of, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Works"

    headers = ["No.", "Project", "Client", "Ref (LPO/LOA)", "Date", "Type",
               "Contract (QAR)", "Received (QAR)", "Balance (QAR)", "Status", "Check"]
    widths = [5, 34, 30, 22, 11, 8, 15, 15, 15, 22, 9]
    _header_block(ws, "APPROVED WORKS REGISTER — ALL AWARDS TO DATE", len(headers), as_of)
    _table_header(ws, 5, headers, widths)

    r = 6
    first_data = r
    tot_contract = tot_received = tot_balance = 0.0

    for i, a in enumerate(awards, start=1):
        contract = float(a["contract"])
        received = float(a["received"])
        balance = contract - received          # computed, never trusted from input
        tot_contract += contract
        tot_received += received
        tot_balance += balance

        vals = [i, a["project"], a["client"], a["ref"], a["date"], a["type"],
                contract, received, balance, a["status"]]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Calibri", size=10)
            c.border = BORDER
            if col in (7, 8, 9):
                c.number_format = MONEY
                c.alignment = Alignment(horizontal="right")
            elif col in (1, 5, 6):
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Status colouring
        st = ws.cell(row=r, column=10)
        s = a["status"].lower()
        if "paid" in s:
            st.fill = PatternFill("solid", fgColor=GREEN_FILL)
            st.font = Font(name="Calibri", size=10, bold=True, color=GREEN_TEXT)
        elif "pending" in s or "blocked" in s:
            st.fill = PatternFill("solid", fgColor=RED_FILL)
            st.font = Font(name="Calibri", size=10, bold=True, color=RED_TEXT)
        else:
            st.fill = PatternFill("solid", fgColor=AMBER)
            st.font = Font(name="Calibri", size=10, bold=True)

        bal = ws.cell(row=r, column=9)
        if balance > 0:
            bal.font = Font(name="Calibri", size=10, bold=True, color=RED_TEXT)

        # Redundant formula cross-check: must agree with the written literal.
        chk = ws.cell(row=r, column=11, value=f"=IF(ROUND(G{r}-H{r}-I{r},2)=0,\"OK\",\"MISMATCH\")")
        chk.font = Font(name="Calibri", size=9, italic=True, color="808080")
        chk.alignment = Alignment(horizontal="center")
        chk.border = BORDER
        r += 1

    last_data = r - 1

    # --- TOTAL row: literal values, plus a formula over the FULL range ---
    ws.cell(row=r, column=1, value="TOTAL")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.border = BORDER
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", indent=1)

    for col, val in ((7, tot_contract), (8, tot_received), (9, tot_balance)):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = MONEY
        c.alignment = Alignment(horizontal="right")
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws.cell(row=r, column=11,
            value=f'=IF(AND(ROUND(SUM(G{first_data}:G{last_data})-G{r},2)=0,'
                  f'ROUND(SUM(H{first_data}:H{last_data})-H{r},2)=0),"OK","MISMATCH")'
            ).font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    total_row = r
    r += 2

    # --- Summary ---
    ws.cell(row=r, column=2, value="SUMMARY").font = Font(name="Calibri", size=12, bold=True,
                                                         color=NAVY)
    r += 1
    for label, val, colour in [
        ("Total Approved Value", tot_contract, NAVY),
        ("Total Received to Date", tot_received, GREEN_TEXT),
        ("Total Outstanding Balance", tot_balance, RED_TEXT),
    ]:
        ws.cell(row=r, column=2, value=label).font = Font(name="Calibri", size=11, bold=True)
        c = ws.cell(row=r, column=7, value=val)
        c.number_format = MONEY
        c.font = Font(name="Calibri", size=11, bold=True, color=colour)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right")
        r += 1

    # --- Concentration: the risk nothing else reports ---
    r += 1
    ws.cell(row=r, column=2, value="CONCENTRATION").font = Font(name="Calibri", size=12,
                                                               bold=True, color=NAVY)
    r += 1
    ranked = sorted(awards, key=lambda a: -float(a["contract"]))
    top2 = sum(float(a["contract"]) for a in ranked[:2])
    pct = (top2 / tot_contract * 100) if tot_contract else 0.0
    for label, val, fmt in [
        (f"Top 2 clients ({ranked[0]['client']}, {ranked[1]['client']})", top2, MONEY),
        ("Top 2 as % of book", pct / 100, '0.0%'),
    ]:
        ws.cell(row=r, column=2, value=label).font = Font(name="Calibri", size=10)
        c = ws.cell(row=r, column=7, value=val)
        c.number_format = fmt
        c.font = Font(name="Calibri", size=11, bold=True,
                      color=RED_TEXT if pct > 60 else NAVY)
        c.fill = PatternFill("solid", fgColor=AMBER if pct > 60 else GREY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right")
        r += 1

    r += 1
    note = ws.cell(row=r, column=1,
                   value="All balances computed as Contract − Received. The 'Check' column "
                         "independently re-derives each row and the total; any cell reading "
                         "MISMATCH means this file has been edited by hand and should be rebuilt.")
    note.font = Font(name="Calibri", size=9, italic=True, color="808080")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A6"
    wb.save(path)
    return {"contract": tot_contract, "received": tot_received,
            "balance": tot_balance, "top2_pct": pct, "total_row": total_row}


def build_receivables(data, as_of, path):
    """Amounts to Receive — next milestone and total remaining kept SEPARATE."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Amounts to Receive"

    headers = ["Project", "Client", "Ref", "Next stage", "Next milestone due (QAR)",
               "Total remaining (QAR)", "Status / blocker"]
    widths = [34, 30, 22, 14, 20, 20, 34]
    _header_block(ws, "AMOUNTS TO RECEIVE", len(headers), as_of)
    _table_header(ws, 5, headers, widths)

    r = 6
    sum_next = sum_total = 0.0
    for item in data:
        nxt = float(item["next_due"])
        rem = float(item["total_remaining"])
        sum_next += nxt
        sum_total += rem
        vals = [item["project"], item["client"], item["ref"], item["stage"], nxt, rem,
                item["status"]]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Calibri", size=10)
            c.border = BORDER
            if col in (5, 6):
                c.number_format = MONEY
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if item.get("blocked"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=RED_FILL)
            ws.cell(row=r, column=7).font = Font(name="Calibri", size=10, bold=True,
                                                 color=RED_TEXT)
        r += 1

    ws.cell(row=r, column=1, value="TOTAL")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.border = BORDER
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", indent=1)
    for col, val in ((5, sum_next), (6, sum_total)):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = MONEY
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="right")

    r += 2
    note = ws.cell(row=r, column=1,
                   value="'Next milestone due' and 'Total remaining' are DIFFERENT quantities "
                         "and must never be summed together. The previous register mixed them "
                         "in a single column, understating what is owed.")
    note.font = Font(name="Calibri", size=9, italic=True, color=RED_TEXT)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A6"
    wb.save(path)
    return {"next": sum_next, "remaining": sum_total}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True)
    ap.add_argument("--outdir", default=".")
    args = ap.parse_args()

    with open(args.data) as f:
        data = json.load(f)

    as_of = data.get("as_of", date.today().strftime("%d %B %Y"))
    os.makedirs(args.outdir, exist_ok=True)

    ap_path = os.path.join(args.outdir, "approved_register.xlsx")
    rc_path = os.path.join(args.outdir, "amounts_to_receive.xlsx")

    a = build_approved(data["awards"], as_of, ap_path)
    b = build_receivables(data["receivables"], as_of, rc_path)

    # Assertions: fail loudly rather than ship a wrong register.
    exp_c = sum(float(x["contract"]) for x in data["awards"])
    exp_r = sum(float(x["received"]) for x in data["awards"])
    assert abs(a["contract"] - exp_c) < 0.01, "contract total mismatch"
    assert abs(a["received"] - exp_r) < 0.01, "received total mismatch"
    assert abs(a["balance"] - (exp_c - exp_r)) < 0.01, "balance total mismatch"

    print(f"Approved Works Register  → {ap_path}")
    print(f"  Total approved value      {a['contract']:>14,.2f}")
    print(f"  Total received to date    {a['received']:>14,.2f}")
    print(f"  Total outstanding         {a['balance']:>14,.2f}")
    print(f"  Top-2 concentration       {a['top2_pct']:>13.1f}%")
    print(f"\nAmounts to Receive       → {rc_path}")
    print(f"  Next milestones due       {b['next']:>14,.2f}")
    print(f"  Total remaining           {b['remaining']:>14,.2f}")
    print("\nAll totals asserted against source data.")


if __name__ == "__main__":
    main()
