#!/usr/bin/env python3
"""Build the yearly maintenance & service log workbook.

    python3 build_maintenance_log.py --data examples/maintenance_clients.json --outdir ./out

Five sheets:

    Clients      one row per covered installation — warranty, AMC, visit frequency
    Service Log  one row per visit — the log itself, everything else reads from it
    Schedule     client × month grid, planned against actually logged
    Summary      per client: coverage, revenue, uninvoiced value, warranty countdown
    Notes        how to run it, and the rules it must not break

The point of the workbook is not record-keeping for its own sake. It answers four
questions nothing in the business currently answers: which warranties are about to
expire (an AMC that never gets proposed), which scheduled visits have been missed,
which completed work was never invoiced, and what maintenance is actually worth per
client per year.

Figures are never invented here. Anything not sourced is left blank and marked, so a
gap looks like a gap instead of a number.
"""

from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# House palette — matches every other TNDK document.
INK = "1F3864"
GOLD = "C9A24E"
TINT = "D6E4F0"
WASH = "F2F2F2"
DUE_TEXT = "C00000"
DUE_FILL = "FDECEA"
OK_FILL = "E2EFDA"
CAUTION = "FFF2CC"

MONEY = '#,##0.00'
DATE_FMT = 'DD-MMM-YY'

# How many Service Log rows the formulas reach across. Generous, because a formula
# range that stops short silently under-reports — the exact failure the Approved
# Works Register is currently making with its 3-row total.
LOG_ROWS = 1000

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

VISIT_TYPES = ["Scheduled AMC", "Warranty call", "Breakdown call-out",
               "Chargeable service", "Inspection", "Installation follow-up"]
STATUSES = ["Completed", "Pending parts", "Follow-up required",
            "Awaiting client", "Cancelled"]

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def title_block(ws, title: str, subtitle: str, ncols: int) -> None:
    ws["A1"] = "THE NEW DOHA KITCHEN EQUIPMENT SERVICES W.L.L."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True, color=INK)
    ws["A2"] = title
    ws["A2"].font = Font(name="Calibri", size=14, bold=True, color=INK)
    ws["A3"] = subtitle
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    for row in (1, 2, 3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(ncols, 8))


def set_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_clients(wb, clients: list[dict], as_of: str):
    ws = wb.create_sheet("Clients")
    headers = ["Client", "Site / Location", "Equipment covered", "Type",
               "Handover date", "Warranty end", "AMC status", "AMC start", "AMC end",
               "AMC value (QAR)", "Visits / year", "Contact", "Source / notes"]
    title_block(ws, "MAINTENANCE CLIENT REGISTER",
                f"One row per covered installation · as of {as_of}", len(headers))

    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row, len(headers))
    # A string reference, not ws.cell(): calling ws.cell() instantiates that
    # cell, which advances max_row and silently pushes every appended data row
    # down by one — leaving this sheet off by one against Summary and Schedule.
    ws.freeze_panes = f"A{header_row + 1}"

    for entry in clients:
        ws.append([
            entry.get("client", ""),
            entry.get("site", ""),
            entry.get("equipment", ""),
            entry.get("type", ""),
            entry.get("handover_date", ""),
            entry.get("warranty_end", ""),
            entry.get("amc_status", "Not offered"),
            entry.get("amc_start", ""),
            entry.get("amc_end", ""),
            entry.get("amc_value", ""),
            entry.get("visits_per_year", ""),
            entry.get("contact", ""),
            entry.get("source", ""),
        ])

    first, last = header_row + 1, ws.max_row
    for row in range(first, last + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (3, 13)))
        for col in (5, 6, 8, 9):
            ws.cell(row=row, column=col).number_format = DATE_FMT
        ws.cell(row=row, column=10).number_format = MONEY

    amc = DataValidation(type="list",
                         formula1='"Not offered,Proposed,Active,Expired,Declined"',
                         allow_blank=True)
    ws.add_data_validation(amc)
    amc.add(f"G{first}:G{last}")

    # Empty warranty dates are the thing most likely to cost money here — an AMC
    # never proposed because nobody knew the cover was ending.
    ws.conditional_formatting.add(
        f"F{first}:F{last}",
        FormulaRule(formula=[f'ISBLANK(F{first})'],
                    fill=PatternFill("solid", fgColor=CAUTION), stopIfTrue=False))
    ws.conditional_formatting.add(
        f"F{first}:F{last}",
        FormulaRule(formula=[f'AND(F{first}<>"",F{first}-TODAY()<=60,F{first}>=TODAY())'],
                    fill=PatternFill("solid", fgColor=DUE_FILL),
                    font=Font(color=DUE_TEXT, bold=True), stopIfTrue=False))

    set_widths(ws, {"A": 32, "B": 20, "C": 26, "D": 14, "E": 13, "F": 13, "G": 14,
                    "H": 12, "I": 12, "J": 15, "K": 12, "L": 22, "M": 46})
    return ws, first, last


def build_service_log(wb, year: int):
    ws = wb.create_sheet("Service Log")
    headers = ["No.", "Date", "Client", "Site", "Equipment", "Visit type",
               "Reported issue / work done", "Parts used", "Technician",
               "Chargeable", "Amount (QAR)", "Invoice ref", "Status",
               "Next visit due", "Remarks"]
    title_block(ws, f"SERVICE LOG {year}",
                "One row per visit. Every other sheet reads from this one — "
                "log the visit here first.", len(headers))

    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row, len(headers))
    ws.freeze_panes = f"C{header_row + 1}"

    first = header_row + 1
    last = first + LOG_ROWS - 1

    for row in range(first, last + 1):
        ws.cell(row=row, column=1, value=f"=IF(B{row}=\"\",\"\",ROW()-{header_row})")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(col in (7, 8, 15)))
        ws.cell(row=row, column=2).number_format = DATE_FMT
        ws.cell(row=row, column=14).number_format = DATE_FMT
        ws.cell(row=row, column=11).number_format = MONEY

    visit = DataValidation(type="list", formula1=f'"{",".join(VISIT_TYPES)}"', allow_blank=True)
    chargeable = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    status = DataValidation(type="list", formula1=f'"{",".join(STATUSES)}"', allow_blank=True)
    for validation, column in ((visit, "F"), (chargeable, "J"), (status, "M")):
        ws.add_data_validation(validation)
        validation.add(f"{column}{first}:{column}{last}")

    # Work done, chargeable, and no invoice number against it. This is money already
    # earned and not billed — the single most valuable thing the log surfaces.
    ws.conditional_formatting.add(
        f"A{first}:O{last}",
        FormulaRule(formula=[f'AND($J{first}="Yes",$L{first}="",$B{first}<>"")'],
                    fill=PatternFill("solid", fgColor=DUE_FILL), stopIfTrue=False))
    # A visit whose next due date has passed with nothing logged since.
    ws.conditional_formatting.add(
        f"N{first}:N{last}",
        CellIsRule(operator="lessThan", formula=["TODAY()"],
                   fill=PatternFill("solid", fgColor=CAUTION)))

    set_widths(ws, {"A": 6, "B": 12, "C": 30, "D": 18, "E": 22, "F": 20, "G": 42,
                    "H": 24, "I": 16, "J": 12, "K": 14, "L": 15, "M": 18,
                    "N": 13, "O": 26})
    return ws, first, last


def build_schedule(wb, clients: list[dict], year: int, log_first: int, log_last: int):
    ws = wb.create_sheet("Schedule")
    headers = ["Client", "Site", "Visits / year"] + MONTHS + ["Logged", "Gap"]
    title_block(ws, f"VISIT SCHEDULE {year}",
                "Counts are read from the Service Log. Gap = visits contracted "
                "minus visits actually logged.", len(headers))

    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row, len(headers))
    ws.freeze_panes = f"C{header_row + 1}"

    first = header_row + 1
    log = "'Service Log'"
    for i, entry in enumerate(clients):
        row = first + i
        client_ref = f"Clients!A{6 + i}"
        ws.cell(row=row, column=1, value=f"={client_ref}")
        ws.cell(row=row, column=2, value=f"=Clients!B{6 + i}")
        ws.cell(row=row, column=3, value=f"=IF(Clients!K{6 + i}=\"\",\"\",Clients!K{6 + i})")

        for m in range(12):
            col = 4 + m
            start = f"DATE({year},{m + 1},1)"
            end = f"DATE({year},{m + 2},1)"
            ws.cell(row=row, column=col, value=(
                f'=COUNTIFS({log}.$C${log_first}:$C${log_last},$A{row},'
                f'{log}.$B${log_first}:$B${log_last},">="&{start},'
                f'{log}.$B${log_first}:$B${log_last},"<"&{end})'
            ).replace(f"{log}.", f"{log}!"))

        logged_col = get_column_letter(16)
        ws.cell(row=row, column=16, value=f"=SUM(D{row}:O{row})")
        ws.cell(row=row, column=17,
                value=f'=IF($C{row}="","",MAX(0,$C{row}-{logged_col}{row}))')

    last = first + len(clients) - 1
    for row in range(first, last + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            if col >= 4:
                cell.alignment = Alignment(horizontal="center")

    ws.conditional_formatting.add(
        f"D{first}:O{last}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=OK_FILL)))
    ws.conditional_formatting.add(
        f"Q{first}:Q{last}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=DUE_FILL),
                   font=Font(color=DUE_TEXT, bold=True)))

    set_widths(ws, {"A": 32, "B": 18, "C": 12, "P": 10, "Q": 8})
    for m in range(12):
        ws.column_dimensions[get_column_letter(4 + m)].width = 5.5
    return ws


def build_summary(wb, clients: list[dict], year: int, log_first: int, log_last: int):
    ws = wb.create_sheet("Summary")
    headers = ["Client", "Visits contracted", "Visits logged", "Coverage",
               "Chargeable revenue (QAR)", "Uninvoiced (QAR)", "Last visit",
               "Next due", "Warranty end", "Days to warranty end", "AMC status",
               "Action"]
    title_block(ws, f"MAINTENANCE SUMMARY {year}",
                "Everything here is computed. Correct the Clients and Service Log "
                "sheets, not this one.", len(headers))

    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row, len(headers))
    ws.freeze_panes = f"B{header_row + 1}"

    first = header_row + 1
    log = "'Service Log'"
    c_first, c_last = f"$C${log_first}", f"$C${log_last}"
    b_first, b_last = f"$B${log_first}", f"$B${log_last}"

    for i, entry in enumerate(clients):
        row = first + i
        cr = 6 + i  # matching row on the Clients sheet
        ws.cell(row=row, column=1, value=f"=Clients!A{cr}")
        ws.cell(row=row, column=2, value=f'=IF(Clients!K{cr}="","",Clients!K{cr})')
        ws.cell(row=row, column=3,
                value=f'=COUNTIFS({log}!{c_first}:{c_last},$A{row},'
                      f'{log}!{b_first}:{b_last},">="&DATE({year},1,1),'
                      f'{log}!{b_first}:{b_last},"<"&DATE({year + 1},1,1))')
        ws.cell(row=row, column=4,
                value=f'=IF(OR($B{row}="",$B{row}=0),"",$C{row}/$B{row})')
        ws.cell(row=row, column=5,
                value=f'=SUMIFS({log}!$K${log_first}:$K${log_last},'
                      f'{log}!{c_first}:{c_last},$A{row},'
                      f'{log}!$J${log_first}:$J${log_last},"Yes")')
        # Chargeable, logged, and no invoice reference — earned and unbilled.
        ws.cell(row=row, column=6,
                value=f'=SUMIFS({log}!$K${log_first}:$K${log_last},'
                      f'{log}!{c_first}:{c_last},$A{row},'
                      f'{log}!$J${log_first}:$J${log_last},"Yes",'
                      f'{log}!$L${log_first}:$L${log_last},"")')
        # MAXIFS / MINIFS rather than MAX(IF(...)): the array form needs
        # Ctrl+Shift+Enter in pre-2019 Excel and returns a wrong single-cell
        # answer without it, which is exactly the kind of quiet error this
        # workbook exists to stop.
        ws.cell(row=row, column=7,
                value=f'=IF($C{row}=0,"",MAXIFS({log}!{b_first}:{b_last},'
                      f'{log}!{c_first}:{c_last},$A{row}))')
        # The latest next-due date, which is the one set at the most recent
        # visit. ">0" rather than "<>" to exclude blanks: dates are serial
        # numbers, and "<>" is read inconsistently across spreadsheet apps.
        ws.cell(row=row, column=8,
                value=f'=IF(COUNTIFS({log}!{c_first}:{c_last},$A{row},'
                      f'{log}!$N${log_first}:$N${log_last},">0")=0,"",'
                      f'MAXIFS({log}!$N${log_first}:$N${log_last},'
                      f'{log}!{c_first}:{c_last},$A{row},'
                      f'{log}!$N${log_first}:$N${log_last},">0"))')
        ws.cell(row=row, column=9, value=f'=IF(Clients!F{cr}="","",Clients!F{cr})')
        ws.cell(row=row, column=10,
                value=f'=IF($I{row}="","",$I{row}-TODAY())')
        ws.cell(row=row, column=11, value=f"=Clients!G{cr}")
        # The one column that says what to do rather than what happened.
        # Every applicable action, not just the first one. Nesting IFs would
        # hide a warranty expiring in three weeks behind an unbilled visit,
        # and both need doing.
        # Plain concatenation with a leading separator on each part, then the
        # first separator trimmed off. TEXTJOIN's ignore-empty flag would read
        # better, but it is honoured inconsistently outside Excel and leaves
        # " · · " where a condition is false.
        parts = (
            f'IF($F{row}>0," · Invoice completed work","")'
            f'&IF(AND($I{row}<>"",$J{row}<=60,$J{row}>=0)," · Warranty ending — propose AMC","")'
            f'&IF(AND($I{row}<>"",$J{row}<0)," · Warranty expired — AMC or lose it","")'
            f'&IF(AND($B{row}<>"",$C{row}<$B{row})," · Visits behind schedule","")'
            f'&IF($I{row}=""," · Confirm warranty end date","")'
        )
        ws.cell(row=row, column=12,
                value=f'=IF(LEN({parts})=0,"",MID({parts},4,300))')

    last = first + len(clients) - 1
    for row in range(first, last + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=4).number_format = '0%'
        ws.cell(row=row, column=5).number_format = MONEY
        ws.cell(row=row, column=6).number_format = MONEY
        ws.cell(row=row, column=7).number_format = DATE_FMT
        ws.cell(row=row, column=8).number_format = DATE_FMT
        ws.cell(row=row, column=9).number_format = DATE_FMT
        ws.cell(row=row, column=12).font = Font(name="Calibri", size=10, bold=True, color=INK)

    total = last + 1
    ws.cell(row=total, column=1, value="TOTAL")
    ws.cell(row=total, column=3, value=f"=SUM(C{first}:C{last})")
    ws.cell(row=total, column=5, value=f"=SUM(E{first}:E{last})")
    ws.cell(row=total, column=6, value=f"=SUM(F{first}:F{last})")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total, column=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.border = BORDER
    ws.cell(row=total, column=5).number_format = MONEY
    ws.cell(row=total, column=6).number_format = MONEY

    ws.conditional_formatting.add(
        f"F{first}:F{last}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=DUE_FILL),
                   font=Font(color=DUE_TEXT, bold=True)))
    ws.conditional_formatting.add(
        f"J{first}:J{last}",
        CellIsRule(operator="between", formula=["0", "60"],
                   fill=PatternFill("solid", fgColor=CAUTION)))

    set_widths(ws, {"A": 32, "B": 15, "C": 13, "D": 10, "E": 20, "F": 16, "G": 12,
                    "H": 12, "I": 13, "J": 18, "K": 14, "L": 32})
    return ws


def build_notes(wb, as_of: str, year: int):
    ws = wb.create_sheet("Notes")
    title_block(ws, "HOW THIS WORKBOOK WORKS", f"Built {as_of}", 2)

    lines = [
        ("", ""),
        ("The one rule", "Log every visit in Service Log on the day it happens. "
                         "Schedule and Summary are formulas — they are only ever as "
                         "current as that sheet."),
        ("", ""),
        ("Clients", "One row per covered installation, not per company. A client with "
                    "a chiller and a freezer on two sites gets two rows, because they "
                    "are serviced and invoiced separately."),
        ("Service Log", f"Every visit in {year}: scheduled AMC, warranty call, "
                        "breakdown, chargeable service. Fill Next visit due before you "
                        "leave site — that is what makes the next one get planned."),
        ("Schedule", "Client against month. Green means a visit was logged. Gap is "
                     "visits contracted minus visits logged."),
        ("Summary", "Computed. Do not type into it — correct Clients or Service Log."),
        ("", ""),
        ("What the colours mean", ""),
        ("Red row in Service Log", "Chargeable work with no invoice reference. Work "
                                   "done and not billed."),
        ("Amber warranty date", "Ends within 60 days — the AMC proposal goes out now, "
                                "not when it lapses."),
        ("Blank warranty date", "Nobody knows when cover ends. Find it in the handover "
                                "record."),
        ("Red gap", "Contracted visits not delivered. If it stays red to year end, the "
                    "AMC was not honoured."),
        ("", ""),
        ("Rules this workbook keeps", ""),
        ("No tax wording", "Nothing here carries a tax or VAT line, and neither does "
                           "any invoice raised from it."),
        ("No invented figures", "A blank is a question, not a zero. An AMC value or "
                                "warranty date that was guessed is worse than an empty "
                                "cell, because nobody checks it again."),
        ("Dates carry an as-of", "Any balance or coverage figure quoted from this "
                                 "workbook is quoted with the date it was read."),
        ("Invoicing stays in LEDGER", "This log flags work to be billed. The invoice "
                                      "itself is raised through the tndk-accounts "
                                      "process, numbered from the numbering log."),
        ("", ""),
        ("Software needed", "Summary uses MAXIFS and MINIFS — Excel 2019, Microsoft 365, "
                            "Google Sheets or LibreOffice 5.2 and above. In an older "
                            "Excel those two columns show #NAME?; everything else still "
                            "works."),
    ]

    for label, body in lines:
        ws.append([label, body])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True, color=INK)
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=row, column=2).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 28 if body else 10

    set_widths(ws, {"A": 26, "B": 88})
    return ws


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("--data", required=True, type=Path)
    parser.add_argument("--outdir", required=True, type=Path)
    parser.add_argument("--year", type=int)
    args = parser.parse_args()

    config = json.loads(args.data.read_text())
    clients = config["clients"]
    year = args.year or config.get("year", date.today().year)
    as_of = config.get("as_of", date.today().strftime("%d %B %Y"))

    wb = Workbook()
    wb.remove(wb.active)

    build_clients(wb, clients, as_of)
    _, log_first, log_last = build_service_log(wb, year)
    build_schedule(wb, clients, year, log_first, log_last)
    build_summary(wb, clients, year, log_first, log_last)
    build_notes(wb, as_of, year)

    wb._sheets = [wb["Summary"], wb["Clients"], wb["Service Log"],
                  wb["Schedule"], wb["Notes"]]

    args.outdir.mkdir(parents=True, exist_ok=True)
    out = args.outdir / f"maintenance_log_{year}.xlsx"
    wb.save(out)

    blanks = sum(1 for c in clients if not c.get("warranty_end"))
    print(f"{out}")
    print(f"  {len(clients)} installations · Service Log rows {log_first}-{log_last}")
    if blanks:
        print(f"  {blanks} of {len(clients)} have no warranty end date — "
              f"amber on the Clients sheet until filled")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
